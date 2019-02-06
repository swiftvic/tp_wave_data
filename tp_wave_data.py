from openpyxl import Workbook
from openpyxl import load_workbook
import datetime


if __name__ == '__main__':

    #wb = Workbook()
    #filepath = "C:/Users/vauyeung/Documents/GitHub/tp_wave_data/demo.xlsx"
    #wb.save(filepath)

    filepath = "demo.xlsx"
    #filepath = "throughput_final.xlsx"

    wb = load_workbook(filepath)

    # Select sheet
    #sheet_ranges = wb['Throughput']

    # Print selected sheet cell value
    #print(sheet_ranges['C7'].value) 

    #sheet_ranges['F11'] = 32

    # Select active sheet
    sheet = wb.active
    #sheet['A1'] = 1

    #sheet.cell(row=2, column = 2).value = "hello"

    wb.save(filepath)